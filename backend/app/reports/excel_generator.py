"""Generate Excel compliance audit reports."""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from app.reports.pdf_generator import extract_ip_from_device


def generate_excel_report(report: dict) -> bytes:
    """Generate Excel bytes from audit report data."""
    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    summary_data = [
        ["Network Compliance Audit Report"],
        [],
        ["Device Name", report.get("device_name", "")],
        ["Device Type", report.get("device_type", "")],
        ["Vendor", report.get("vendor", "Cisco")],
        ["Audit Mode", report.get("audit_mode", "full")],
        ["Overall Score", f"{report.get('overall_score', 0)}%"],
        ["Generated", str(report.get("created_at", datetime.utcnow().isoformat()))],
    ]
    for row in summary_data:
        ws_summary.append(row)

    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 40

    ws_categories = wb.create_sheet("Category Scores")
    ws_categories.append(["Category", "Score (%)"])
    for cell in ws_categories[1]:
        cell.fill = header_fill
        cell.font = header_font

    for cat, score in sorted(report.get("category_scores", {}).items()):
        ws_categories.append([cat.replace("_", " ").title(), score])
    ws_categories.column_dimensions["A"].width = 25
    ws_categories.column_dimensions["B"].width = 15

    ws_passed = wb.create_sheet("Passed Rules")
    ws_passed.append(["Rule", "Category", "Status"])
    for cell in ws_passed[1]:
        cell.fill = PatternFill(start_color="16A34A", end_color="16A34A", fill_type="solid")
        cell.font = header_font
    for rule in report.get("passed", []):
        ws_passed.append([rule.get("rule", ""), rule.get("category", ""), rule.get("status", "PASS")])
    ws_passed.column_dimensions["A"].width = 60
    ws_passed.column_dimensions["B"].width = 15

    ws_failed = wb.create_sheet("Failed Rules")
    ws_failed.append(["Rule", "Category", "Status"])
    for cell in ws_failed[1]:
        cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        cell.font = header_font
    for rule in report.get("failed", []):
        ws_failed.append([rule.get("rule", ""), rule.get("category", ""), rule.get("status", "FAIL")])
    ws_failed.column_dimensions["A"].width = 60

    ws_recs = wb.create_sheet("Recommendations")
    ws_recs.append(["Rule", "Recommendation", "Remediation Command"])
    for cell in ws_recs[1]:
        cell.fill = header_fill
        cell.font = header_font
    for rec in report.get("failed", []):
        if rec.get("recommendation"):
            ws_recs.append([
                rec.get("rule", ""),
                rec.get("recommendation", ""),
                rec.get("remediation", "")
            ])
    ws_recs.column_dimensions["A"].width = 40
    ws_recs.column_dimensions["B"].width = 50
    ws_recs.column_dimensions["C"].width = 50

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def generate_group_excel_report(group_info: dict, devices: list, audit_results: list) -> bytes:
    """Generate consolidated Excel workbook report for a device group."""
    wb = Workbook()

    # Sheet 1: Group Summary
    ws_summary = wb.active
    ws_summary.title = "Group Summary"

    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    dark_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    # Title
    ws_summary.append(["Device Group Compliance Report"])
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.append([])

    # Group Info Block
    ws_summary.append(["Group Information"])
    ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(bold=True)
    
    ws_summary.append(["Vendor", group_info.get("vendor", "N/A")])
    ws_summary.append(["Device Type", group_info.get("device_type", "N/A")])
    ws_summary.append(["Model", group_info.get("model", "N/A")])
    ws_summary.append(["Template Name", group_info.get("template_name", "N/A") or "N/A"])
    ws_summary.append(["Audit Mode", str(group_info.get("audit_mode", "full")).title()])
    if group_info.get("audit_mode") == "sections":
        ws_summary.append(["Selected Sections", ", ".join(group_info.get("selected_sections", []))])
    
    ws_summary.append(["Number of Devices", len(devices)])
    ws_summary.append([])

    # Calculations
    passed_count = 0
    failed_count = 0
    total_score = 0.0
    valid_results_count = 0
    
    device_results = {r["device_id"]: r for r in audit_results if r}
    
    for d in devices:
        res = device_results.get(str(d["_id"]))
        if res:
            score = res.get("overall_score", 0.0)
            total_score += score
            valid_results_count += 1
            if score >= 80:
                passed_count += 1
            else:
                failed_count += 1
        else:
            failed_count += 1
            
    compliance_pct = (total_score / valid_results_count) if valid_results_count > 0 else 0.0

    # Executive Summary Table
    ws_summary.append(["Executive Summary"])
    ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(bold=True)
    
    ws_summary.append(["Total Devices", "Passed Devices", "Failed Devices", "Compliance Percentage"])
    for col_idx in range(1, 5):
        cell = ws_summary.cell(row=ws_summary.max_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    ws_summary.append([len(devices), passed_count, failed_count, f"{compliance_pct:.2f}%"])
    for col_idx in range(1, 5):
        ws_summary.cell(row=ws_summary.max_row, column=col_idx).alignment = Alignment(horizontal="center")
        
    ws_summary.append([])

    # Device Summary list
    ws_summary.append(["Device Summary Table"])
    ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(bold=True)
    
    ws_summary.append(["Device Name", "IP Address", "Compliance Score", "Audit Result"])
    for col_idx in range(1, 5):
        cell = ws_summary.cell(row=ws_summary.max_row, column=col_idx)
        cell.fill = dark_fill
        cell.font = header_font
        
    for d in devices:
        dev_id = str(d["_id"])
        res = device_results.get(dev_id)
        ip = extract_ip_from_device(d)
        score_val = "—"
        result_val = "PENDING"
        if res:
            score_val = f"{round(res.get('overall_score', 0))}%"
            result_val = "PASS" if res.get('overall_score', 0) >= 80 else "FAIL"
            
        ws_summary.append([
            d.get("device_name", "N/A"),
            ip,
            score_val,
            result_val
        ])

    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 20
    ws_summary.column_dimensions["C"].width = 20
    ws_summary.column_dimensions["D"].width = 15

    # Sheet 2: Device Details
    ws_details = wb.create_sheet("Device Details")
    ws_details.append(["Device Name", "IP Address", "Rule", "Category", "Status", "Recommendation", "Remediation"])
    for cell in ws_details[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    for d in devices:
        dev_id = str(d["_id"])
        res = device_results.get(dev_id)
        ip = extract_ip_from_device(d)
        
        if res:
            for rule in res.get("passed", []):
                ws_details.append([
                    d.get("device_name"),
                    ip,
                    rule.get("rule"),
                    rule.get("category"),
                    "PASS",
                    "",
                    ""
                ])
            for rule in res.get("failed", []):
                ws_details.append([
                    d.get("device_name"),
                    ip,
                    rule.get("rule"),
                    rule.get("category"),
                    "FAIL",
                    rule.get("recommendation", ""),
                    rule.get("remediation", "")
                ])
        else:
            ws_details.append([
                d.get("device_name"),
                ip,
                "No audit results available",
                "",
                "PENDING",
                "",
                ""
            ])
            
    ws_details.column_dimensions["A"].width = 20
    ws_details.column_dimensions["B"].width = 15
    ws_details.column_dimensions["C"].width = 40
    ws_details.column_dimensions["D"].width = 15
    ws_details.column_dimensions["E"].width = 10
    ws_details.column_dimensions["F"].width = 40
    ws_details.column_dimensions["G"].width = 40

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def generate_upload_excel_report(upload_info: dict, groups: list, devices_by_group: dict, audit_results_by_device: dict) -> bytes:
    """Generate consolidated Excel workbook report for an entire upload."""
    wb = Workbook()

    # Sheet 1: Upload Summary
    ws_summary = wb.active
    ws_summary.title = "Upload Summary"

    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    dark_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws_summary.append(["Complete Upload Compliance Audit Report"])
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.append([])

    # Metadata Block
    created_at_str = str(upload_info.get("created_at", ""))
    if isinstance(upload_info.get("created_at"), datetime):
        created_at_str = upload_info["created_at"].strftime("%Y-%m-%d %H:%M:%S UTC")

    ws_summary.append(["Upload Details"])
    ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(bold=True)
    ws_summary.append(["Folder Name", upload_info.get("folder_name", "N/A")])
    ws_summary.append(["Upload Date", created_at_str])
    ws_summary.append(["Uploaded By", upload_info.get("created_by", "N/A") or "N/A"])
    ws_summary.append(["Total Groups", len(groups)])
    ws_summary.append(["Total Devices", upload_info.get("total_devices", 0)])
    ws_summary.append([])

    # Overall and Group Calculations
    overall_passed_devices = 0
    overall_failed_devices = 0
    overall_total_score = 0.0
    overall_valid_results_count = 0
    
    group_summaries = []
    
    for group in groups:
        g_id = group["group_id"]
        g_devices = devices_by_group.get(g_id, [])
        
        g_passed_count = 0
        g_failed_count = 0
        g_total_score = 0.0
        g_valid_count = 0
        
        for d in g_devices:
            dev_id = str(d["_id"])
            res = audit_results_by_device.get(dev_id)
            if res:
                score = res.get("overall_score", 0.0)
                g_total_score += score
                g_valid_count += 1
                
                if score >= 80:
                    g_passed_count += 1
                    overall_passed_devices += 1
                else:
                    g_failed_count += 1
                    overall_failed_devices += 1
            else:
                g_failed_count += 1
                overall_failed_devices += 1
                
        g_compliance = (g_total_score / g_valid_count) if g_valid_count > 0 else 0.0
        
        overall_total_score += g_total_score
        overall_valid_results_count += g_valid_count
        
        group_summaries.append({
            "vendor": group.get("vendor", "N/A"),
            "device_type": group.get("device_type", "N/A"),
            "model": group.get("model", "N/A"),
            "device_count": len(g_devices),
            "compliance": g_compliance
        })
        
    overall_compliance_pct = (overall_total_score / overall_valid_results_count) if overall_valid_results_count > 0 else 0.0

    # Executive Summary Table
    ws_summary.append(["Overall Compliance Summary"])
    ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(bold=True)
    
    ws_summary.append(["Total Devices Audited", "Passed Devices", "Failed Devices", "Overall Compliance Score"])
    for col_idx in range(1, 5):
        cell = ws_summary.cell(row=ws_summary.max_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    ws_summary.append([overall_valid_results_count, overall_passed_devices, overall_failed_devices, f"{overall_compliance_pct:.2f}%"])
    for col_idx in range(1, 5):
        ws_summary.cell(row=ws_summary.max_row, column=col_idx).alignment = Alignment(horizontal="center")
        
    ws_summary.append([])

    # Group Summary Table
    ws_summary.append(["Group Summaries"])
    ws_summary.cell(row=ws_summary.max_row, column=1).font = Font(bold=True)
    
    ws_summary.append(["Vendor", "Device Type", "Model", "Device Count", "Compliance Score"])
    for col_idx in range(1, 6):
        cell = ws_summary.cell(row=ws_summary.max_row, column=col_idx)
        cell.fill = dark_fill
        cell.font = header_font
        
    for gs in group_summaries:
        ws_summary.append([
            gs["vendor"],
            gs["device_type"],
            gs["model"],
            gs["device_count"],
            f"{gs['compliance']:.2f}%"
        ])
        
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 15
    ws_summary.column_dimensions["D"].width = 15
    ws_summary.column_dimensions["E"].width = 20

    # Sheet 2: Devices Compliance
    ws_devices = wb.create_sheet("Devices Compliance")
    ws_devices.append(["Group ID", "Device Name", "IP Address", "Compliance Score", "Audit Result"])
    for cell in ws_devices[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    for group in groups:
        g_id = group["group_id"]
        g_devices = devices_by_group.get(g_id, [])
        for d in g_devices:
            dev_id = str(d["_id"])
            res = audit_results_by_device.get(dev_id)
            ip = extract_ip_from_device(d)
            score_val = "—"
            result_val = "PENDING"
            if res:
                score_val = f"{round(res.get('overall_score', 0))}%"
                result_val = "PASS" if res.get('overall_score', 0) >= 80 else "FAIL"
                
            ws_devices.append([
                g_id,
                d.get("device_name"),
                ip,
                score_val,
                result_val
            ])
            
    ws_devices.column_dimensions["A"].width = 30
    ws_devices.column_dimensions["B"].width = 25
    ws_devices.column_dimensions["C"].width = 20
    ws_devices.column_dimensions["D"].width = 20
    ws_devices.column_dimensions["E"].width = 15

    # Sheet 3: All Failed Rules
    ws_failures = wb.create_sheet("All Failed Rules")
    ws_failures.append(["Group ID", "Device Name", "IP Address", "Rule", "Category", "Recommendation", "Remediation"])
    for cell in ws_failures[1]:
        cell.fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        cell.font = header_font
        
    for group in groups:
        g_id = group["group_id"]
        g_devices = devices_by_group.get(g_id, [])
        for d in g_devices:
            dev_id = str(d["_id"])
            res = audit_results_by_device.get(dev_id)
            ip = extract_ip_from_device(d)
            if res:
                for rule in res.get("failed", []):
                    ws_failures.append([
                        g_id,
                        d.get("device_name"),
                        ip,
                        rule.get("rule"),
                        rule.get("category"),
                        rule.get("recommendation", ""),
                        rule.get("remediation", "")
                    ])
                    
    ws_failures.column_dimensions["A"].width = 30
    ws_failures.column_dimensions["B"].width = 25
    ws_failures.column_dimensions["C"].width = 20
    ws_failures.column_dimensions["D"].width = 40
    ws_failures.column_dimensions["E"].width = 15
    ws_failures.column_dimensions["F"].width = 40
    ws_failures.column_dimensions["G"].width = 40

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
